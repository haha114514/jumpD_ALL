import os
import random
import time
import json
import requests
from openpyxl import load_workbook
from hoshino import util, R
from hoshino import Service
import aiohttp
from PIL import Image
from io import BytesIO
from os import path
plugin_path = '/root/HoshinoBot/hoshino/modules/jumpd-all/' #插件位置
res_path = '/root/HoshinoBot/res/img/' #保存资源位置
file = plugin_path +'all.xlsx' #歌曲资料表单位置
un_use_str = '123.'

#随机从表格中选择一行，并生成json
class ExcelUtils:
    def __init__(self):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    # 行数
    def get_rows(self):
        rows = self.ws.max_row
        return rows

    # 列数
    def get_clos(self):
        clo = self.ws.max_column
        return clo

    # 获取值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 修改值并保存
    def set_cell_value(self, row, column, cell_value):
        try:
            self.ws.cell(row=row, column=column).value = cell_value
            self.wb.save(self.file)
        except Exception as e:
            print("error :{}".format(e))
            self.wb.save(self.file)

    # 替换单元格中的内容
    def replace_cell_value(self):
        # 遍历第一行的值，
        for i in range(1, self.get_clos() + 1):
            cell_value = self.get_cell_value(1, i)
            # 是否存在需要替换的值
            if un_use_str in cell_value:
                cell_replace = cell_value.replace(un_use_str, "")
                self.set_cell_value(1, i, cell_replace)


def to_json():
    excel_utils = ExcelUtils()
    excel_dict = {}

    clo = excel_utils.get_clos()
    # 遍历excel中的值存入字典中
    num = random.randint(2, 30)
    for i in range(1, clo + 1):

        dict_key = excel_utils.get_cell_value(1, i)
        dict_value = excel_utils.get_cell_value(num, i) #读取指定行内容
        excel_dict[dict_key] = dict_value
    # 字典转json
    excel_json = json.dumps(excel_dict)
    return excel_json


sv = Service('jumpD', visible= False, enable_on_default= True, bundle='jumpD', help_='''
投硬币
'''.strip())
@sv.on_fullmatch(('jumpD'))
async def jumpD(bot, ev):
    #从json中读取数据
    compare_json = to_json().encode('utf-8').decode('unicode_escape')
    #print(compare_json)
    song_info = json.loads(compare_json)
    Category = song_info["Category"]
    Song = song_info["Song"]
    Artist = song_info["Artist"]
    BPM = song_info["BPM"]
    easy = song_info["easy"]
    easycombo = song_info["easycombo"]
    normal = song_info["normal"]
    normalcombo = song_info["normalcombo"]
    video = song_info["video"]
    notice = song_info["notice"]
    origin = song_info["origin"]
    url = song_info["cover"]
    #下载url里的图片，并保存到res文件夹下的temp.jpg(注意：由于图片下载链接为Google，请部署本插件到海外服务器上)
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as resp:
            cont = await resp.read()
            img = Image.open(BytesIO(cont))
            img.save(path.join(path.dirname(res_path), 'temp.jpg'))
            picfile = path.join(path.dirname(res_path), 'temp.jpg')
            fuckimg = R.img('temp.jpg').cqcode
            msg = f"我觉得你可以跳\n{fuckimg}"+ "\n歌曲:" + str(Song) + "\n艺人:" + str(Artist) + "\nBPM:" + str(BPM) + "\n类别:"  + str(Category) + "\nEasy Lv" + str(easy) + " Note:" + str(easycombo)  + "\nNormal Lv" + str(normal) + " Note:" + str(normalcombo) + "\n出处:"+ str(origin) + "\nPremium Mode:"+ str(video)  + "\nPS:"+ str(notice)
            print(msg)
            await bot.send(ev,msg)

          